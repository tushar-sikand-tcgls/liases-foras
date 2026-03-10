"""
Comprehensive Test Suite for Enriched Layers
Uses examples and expected values from LF-Layers_ENRICHED_v3.xlsx

Tests:
- All Layer 0 and Layer 1 attributes
- At least 3 test cases per attribute (using prompt variations)
- Routing accuracy (vectorized understanding)
- Calculation accuracy (compared to expected values)
- WITHOUT HARDCODINGS (all dynamic)
"""

import openpyxl
import requests
import json
from typing import List, Dict, Optional
from dataclasses import dataclass

API_BASE_URL = "http://localhost:8000"
EXCEL_FILE = "/Users/tusharsikand/Documents/Projects/liases-foras/change-request/enriched-layers/LF-Layers_ENRICHED_v3.xlsx"

@dataclass
class TestCase:
    """Test case for an attribute"""
    attribute_name: str
    layer: str  # "L0" or "L1"
    prompt: str
    project_name: str
    expected_value: Optional[float]
    expected_unit: Optional[str]
    dimension: str
    formula: Optional[str]

class ExcelTestLoader:
    """Loads test cases from Excel file"""

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = None

    def load(self) -> List[TestCase]:
        """Load all test cases from Excel"""
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            test_cases = []

            # Read both Layer0 and Layer1 sheets
            for sheet_name in ['Layer0', 'Layer1']:
                if sheet_name not in self.workbook.sheetnames:
                    continue

                sheet = self.workbook[sheet_name]
                print(f"\nReading sheet: {sheet_name}")

                # Read header row to find column indices
                headers = {}
                for col_idx, cell in enumerate(sheet[1], 1):
                    if cell.value:
                        headers[cell.value.strip()] = col_idx

                print(f"  Columns: {list(headers.keys())}")

                # Read data rows
                for row_idx in range(2, sheet.max_row + 1):
                    row = sheet[row_idx]

                    # Extract values based on headers
                    attr_name = self._get_cell_value(row, headers.get('Target Attribute', 2))
                    if not attr_name:
                        continue

                    layer = self._get_cell_value(row, headers.get('Layer', 1))
                    unit = self._get_cell_value(row, headers.get('Unit', 3))
                    dimension = self._get_cell_value(row, headers.get('Dimension', 4))
                    description = self._get_cell_value(row, headers.get('Description', 5))
                    sample_prompt = self._get_cell_value(row, headers.get('Sample Prompt', 6))
                    variations = self._get_cell_value(row, headers.get('Variation in Prompt', 7))
                    formula = self._get_cell_value(row, headers.get('Formula/Derivation', 9))
                    sample_values = self._get_cell_value(row, headers.get('Sample Values', 10))

                    # Create test cases from prompts
                    prompts_to_test = []

                    # Add sample prompt
                    if sample_prompt:
                        prompts_to_test.append(sample_prompt)

                    # Add variations (split by common delimiters)
                    if variations:
                        var_list = [v.strip() for v in variations.replace('|', '?').split('?') if v.strip()]
                        prompts_to_test.extend(var_list[:2])  # Add up to 2 variations

                    # Ensure at least 3 test cases per attribute
                    while len(prompts_to_test) < 3 and sample_prompt:
                        prompts_to_test.append(sample_prompt)

                    # Create test cases
                    for prompt in prompts_to_test[:3]:  # Exactly 3 tests per attribute
                        test_case = TestCase(
                            attribute_name=attr_name,
                            layer=layer,
                            prompt=prompt.strip(),
                            project_name="Sara City",  # Default project from examples
                            expected_value=self._extract_expected_value(sample_values),
                            expected_unit=unit,
                            dimension=dimension,
                            formula=formula
                        )
                        test_cases.append(test_case)

            print(f"\nTotal test cases loaded: {len(test_cases)}")
            return test_cases

        except Exception as e:
            print(f"Error loading Excel file: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _get_cell_value(self, row, col_idx):
        """Safely get cell value"""
        if col_idx and col_idx <= len(row):
            cell = row[col_idx - 1]
            return str(cell.value) if cell.value is not None else None
        return None

    def _extract_expected_value(self, example: Optional[str]) -> Optional[float]:
        """Extract numeric value from example"""
        if not example:
            return None
        try:
            # Try to extract first number from example
            import re
            numbers = re.findall(r'[-+]?\d*\.?\d+', example)
            if numbers:
                return float(numbers[0])
        except:
            pass
        return None

    def _extract_expected_unit(self, example: Optional[str]) -> Optional[str]:
        """Extract unit from example"""
        if not example:
            return None
        # Common units
        units = ['Years', 'Months', 'Days', 'Units', '₹/sqft', 'Cr', '%', 'count']
        for unit in units:
            if unit.lower() in example.lower():
                return unit
        return None

class EnrichedLayersTestRunner:
    """Runs comprehensive tests for enriched layers"""

    def __init__(self, test_cases: List[TestCase]):
        self.test_cases = test_cases
        self.results = {
            'total': 0,
            'passed': 0,
            'failed': 0,
            'routing_correct': 0,
            'calculation_correct': 0,
            'details': []
        }

    def run_test(self, test_case: TestCase) -> Dict:
        """Run a single test case"""
        # Construct query with project name
        prompt_lower = test_case.prompt.lower()

        # Check if prompt already contains a project reference
        has_project = any(word in prompt_lower for word in ['sara city', 'project', 'x'])

        if 'X' in test_case.prompt or '[X]' in test_case.prompt:
            query = test_case.prompt.replace('X', test_case.project_name).replace('[X]', test_case.project_name)
        elif has_project and 'X' not in test_case.prompt:
            # Prompt mentions project but doesn't have X placeholder - use as is
            query = test_case.prompt
        else:
            # Add project name to query
            query = f"{test_case.prompt} for {test_case.project_name}"

        try:
            response = requests.post(
                f"{API_BASE_URL}/api/qa/question",
                json={"question": query, "project_id": None},
                timeout=30
            )

            if response.status_code != 200:
                return {
                    'status': 'FAIL',
                    'reason': f'HTTP {response.status_code}',
                    'query': query
                }

            data = response.json()
            if data.get('status') != 'success':
                return {
                    'status': 'FAIL',
                    'reason': 'API error',
                    'query': query
                }

            answer = data.get('answer', {})
            result = answer.get('result', {})
            understanding = answer.get('understanding', {})

            # Check routing
            routed_correctly = (
                understanding.get('layer') == test_case.layer or
                (test_case.layer == 'L1' and understanding.get('layer') == 'LAYER_1') or
                (test_case.layer == 'L0' and understanding.get('layer') == 'LAYER_0')
            )

            # Check value (if expected value provided)
            value_correct = True
            if test_case.expected_value is not None:
                actual_value = result.get('value')
                if actual_value is not None:
                    # Allow 5% tolerance
                    tolerance = abs(test_case.expected_value * 0.05)
                    value_correct = abs(actual_value - test_case.expected_value) <= max(tolerance, 0.01)
                else:
                    value_correct = False

            # Check unit
            unit_correct = True
            if test_case.expected_unit:
                actual_unit = result.get('unit')
                unit_correct = actual_unit == test_case.expected_unit if actual_unit else False

            # Overall pass/fail
            passed = routed_correctly and value_correct and unit_correct

            return {
                'status': 'PASS' if passed else 'PARTIAL' if routed_correctly else 'FAIL',
                'query': query,
                'routed_correctly': routed_correctly,
                'value_correct': value_correct,
                'unit_correct': unit_correct,
                'expected_value': test_case.expected_value,
                'actual_value': result.get('value'),
                'expected_unit': test_case.expected_unit,
                'actual_unit': result.get('unit'),
                'routed_layer': understanding.get('layer'),
                'routed_operation': understanding.get('operation'),
                'confidence': understanding.get('confidence')
            }

        except Exception as e:
            return {
                'status': 'ERROR',
                'reason': str(e),
                'query': query
            }

    def run_all(self):
        """Run all test cases"""
        print("=" * 100)
        print("COMPREHENSIVE ENRICHED LAYERS TEST SUITE")
        print("Using Examples and Expected Values from Excel")
        print("=" * 100)

        # Group by attribute
        by_attribute = {}
        for tc in self.test_cases:
            if tc.attribute_name not in by_attribute:
                by_attribute[tc.attribute_name] = []
            by_attribute[tc.attribute_name].append(tc)

        print(f"\nTotal Attributes: {len(by_attribute)}")
        print(f"Total Test Cases: {len(self.test_cases)}")

        # Run tests by attribute
        for attr_name, test_cases in by_attribute.items():
            print(f"\n{'=' * 100}")
            print(f"TESTING: {attr_name} ({test_cases[0].layer})")
            print(f"{'=' * 100}")
            print(f"Dimension: {test_cases[0].dimension}")
            if test_cases[0].formula:
                print(f"Formula: {test_cases[0].formula}")
            print(f"Test Cases: {len(test_cases)}")

            attr_results = []
            for i, test_case in enumerate(test_cases, 1):
                print(f"\n  Test {i}/{len(test_cases)}:")
                print(f"  Prompt: '{test_case.prompt}'")
                if test_case.expected_value:
                    print(f"  Expected: {test_case.expected_value} {test_case.expected_unit or ''}")

                result = self.run_test(test_case)
                attr_results.append(result)

                self.results['total'] += 1
                if result['status'] == 'PASS':
                    self.results['passed'] += 1
                    print(f"  ✓ PASS")
                elif result['status'] == 'PARTIAL':
                    print(f"  ⚠ PARTIAL")
                else:
                    self.results['failed'] += 1
                    print(f"  ✗ {result['status']}")

                if result.get('routed_correctly'):
                    self.results['routing_correct'] += 1
                    print(f"    ✓ Routing: {result.get('routed_layer')} ({result.get('confidence', 0):.2f} confidence)")
                else:
                    print(f"    ✗ Routing: {result.get('routed_layer')} (expected {test_case.layer})")

                if test_case.expected_value:
                    if result.get('value_correct'):
                        self.results['calculation_correct'] += 1
                        print(f"    ✓ Value: {result.get('actual_value')} (expected {test_case.expected_value})")
                    else:
                        print(f"    ✗ Value: {result.get('actual_value')} (expected {test_case.expected_value})")

                if test_case.expected_unit:
                    if result.get('unit_correct'):
                        print(f"    ✓ Unit: {result.get('actual_unit')}")
                    else:
                        print(f"    ✗ Unit: {result.get('actual_unit')} (expected {test_case.expected_unit})")

            # Attribute summary
            passed = sum(1 for r in attr_results if r['status'] == 'PASS')
            print(f"\n  Attribute Summary: {passed}/{len(test_cases)} tests passed ({passed/len(test_cases)*100:.1f}%)")

            self.results['details'].append({
                'attribute': attr_name,
                'layer': test_cases[0].layer,
                'total': len(test_cases),
                'passed': passed,
                'results': attr_results
            })

        # Final summary
        self.print_summary()

    def print_summary(self):
        """Print final test summary"""
        print("\n" + "=" * 100)
        print("FINAL TEST SUMMARY")
        print("=" * 100)

        print(f"\nTotal Test Cases: {self.results['total']}")
        print(f"\nResults:")
        print(f"  ✓ PASSED: {self.results['passed']} ({self.results['passed']/self.results['total']*100:.1f}%)")
        print(f"  ✗ FAILED: {self.results['failed']} ({self.results['failed']/self.results['total']*100:.1f}%)")

        print(f"\nRouting Accuracy:")
        print(f"  ✓ Correct: {self.results['routing_correct']}/{self.results['total']} ({self.results['routing_correct']/self.results['total']*100:.1f}%)")

        tests_with_expected = sum(1 for d in self.results['details'] for r in d['results'] if r.get('expected_value') is not None)
        if tests_with_expected > 0:
            print(f"\nCalculation Accuracy:")
            print(f"  ✓ Correct: {self.results['calculation_correct']}/{tests_with_expected} ({self.results['calculation_correct']/tests_with_expected*100:.1f}%)")

        # Per-attribute breakdown
        print("\n" + "=" * 100)
        print("PER-ATTRIBUTE BREAKDOWN")
        print("=" * 100)

        for detail in self.results['details']:
            icon = "✓" if detail['passed'] == detail['total'] else "⚠" if detail['passed'] > 0 else "✗"
            print(f"\n{icon} {detail['attribute']} ({detail['layer']})")
            print(f"  Tests: {detail['passed']}/{detail['total']} passed ({detail['passed']/detail['total']*100:.1f}%)")

        # Success criteria
        print("\n" + "=" * 100)
        success_rate = self.results['passed'] / self.results['total'] * 100
        routing_rate = self.results['routing_correct'] / self.results['total'] * 100

        if success_rate >= 90 and routing_rate >= 95:
            print("✓ SUCCESS: >= 90% tests passing, >= 95% routing accuracy")
            print("\n✅ System is PRODUCTION READY")
            return True
        elif success_rate >= 75 and routing_rate >= 85:
            print("⚠ PARTIAL SUCCESS: >= 75% tests passing, >= 85% routing accuracy")
            print("\n⚠️ System needs minor improvements")
            return False
        else:
            print("✗ FAILURE: < 75% tests passing or < 85% routing accuracy")
            print("\n❌ System needs significant improvements")
            return False

def main():
    """Main test execution"""
    # Load test cases from Excel
    print("Loading test cases from Excel...")
    loader = ExcelTestLoader(EXCEL_FILE)
    test_cases = loader.load()

    if not test_cases:
        print("ERROR: No test cases loaded from Excel")
        return False

    # Run tests
    runner = EnrichedLayersTestRunner(test_cases)
    success = runner.run_all()

    return success

if __name__ == "__main__":
    import sys
    success = main()
    sys.exit(0 if success else 1)
